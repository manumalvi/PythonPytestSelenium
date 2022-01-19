import openpyxl


class TestData:
    test_Homepage_Data = [{"firstname" : "manu","lastname" : "malvi","gender" : "Male"}, {"firstname" : "honey","lastname" : "malviya","gender" : "Female"}]

# better to use data indict form only excel is old methdo
    @staticmethod   # we can call method of any class by not creating object,self will be removed
    def getTestDataExcel(test_case_name):
        book = openpyxl.load_workbook("C:\\Users\\mmalvi\\Desktop\\pythonexcel.xlsx")
        Dict = {}
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:  # it will print data of column Test2
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = (sheet.cell(row=i, column=j)).value
        return[Dict]

